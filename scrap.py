import pandas as pd
import re

def preprocess_wo_number(df):
    input_col = df["REQ_WO_NUM"].astype(str)
    df_output = pd.DataFrame()
    df_output["wo_original"] = input_col

    # Remove dashes or dots based on segment length
    df_output["removed_dashes"] = df_output["wo_original"].apply(
        lambda x: x.replace(",", "").replace("-", " ") if len(x.replace(",", "").split("-")[0]) < 9 
        else x.replace(",", "").replace("-", "")
    )

    # Extract work order numbers using regex
    pattern = r"TD[ ]?\d{6,7}|90\d{7,8}|80\d{7,8}"
    df_output["all_wo"] = df_output["removed_dashes"].str.upper().apply(
        lambda x: re.findall(pattern, x)
    )

    return df_output

df_sap_current["ReceiverWorkOrderID"] = (
    df_sap_current["ReceiverWorkOrderID"]
    .str.lstrip("0")
    .str.replace(r"\.0*", "", regex=True)
)



import pandas as pd
import re

# ------------------------------
# 1. Sample iTOA Data
# ------------------------------
itoa_data = pd.DataFrame({
    "APP_NUM": [1, 2, 3, 4, 5],
    "REQ_WO_NUM": [
        "TD11111",                  
        "TD22222, TD33333",         
        "TD44444",                  
        "TD55555",                  
        "TD66666, TD77777, TD88888"
    ],
    "REQ_SCHED_START_DATE": pd.to_datetime([
        "2025-04-06", "2025-04-07", "2025-03-01", "2025-04-10", "2025-04-05"
    ]),
    "REQ_SCHED_END_DATE": pd.to_datetime([
        "2025-04-08", "2025-04-09", "2025-04-10", "2025-04-12", "2025-04-06"
    ])
})

# ------------------------------
# 2. Sample SAP Data
# ------------------------------
sap_data = pd.DataFrame({
    "Work Order #": ["TD11111", "TD33333", "TD55555", "TD88888", "TD44444"],
    "Date": pd.to_datetime([
        "2025-04-07", "2025-04-08", "2025-04-11", "2025-04-06", "2025-04-10"
    ])
})

# ------------------------------
# 3. Preprocess iTOA WO numbers
# ------------------------------
def preprocess_wo_number(df):
    input_col = df["REQ_WO_NUM"].astype(str)
    df_output = pd.DataFrame()
    df_output["wo_original"] = input_col
    df_output["APP_NUM"] = df["APP_NUM"]

    df_output["removed_dashes"] = df_output["wo_original"].apply(
        lambda x: x.replace(",", "").replace("-", " ") if len(x.replace(",", "").split("-")[0]) < 9 
        else x.replace(",", "").replace("-", "")
    )

    pattern = r"TD[ ]?\d{5,7}|90\d{7,8}|80\d{7,8}"  # Accepts 5–7 digits
    df_output["all_wo"] = df_output["removed_dashes"].str.upper().apply(
        lambda x: re.findall(pattern, x)
    )

    return df_output

# ------------------------------
# 4. Analysis Date Range
# ------------------------------
analysis_start = pd.to_datetime("2025-04-06")
analysis_end = pd.to_datetime("2025-04-12")

# Filter iTOA to those with start or end within range
mask = (
    (itoa_data["REQ_SCHED_START_DATE"].between(analysis_start, analysis_end)) |
    (itoa_data["REQ_SCHED_END_DATE"].between(analysis_start, analysis_end))
)
itoa_filtered = itoa_data[mask].copy()

# ------------------------------
# 5. Explode Work Orders
# ------------------------------
df_itoa_processed = preprocess_wo_number(itoa_filtered)
df_exploded = df_itoa_processed.explode("all_wo").drop_duplicates()
df_exploded = df_exploded.rename(columns={"all_wo": "REQ_WO_NUM_CLEAN"})

# Add REQ_SCHED_* columns back
df_exploded = df_exploded.merge(
    itoa_filtered[["APP_NUM", "REQ_SCHED_START_DATE", "REQ_SCHED_END_DATE"]],
    on="APP_NUM", how="left"
)

# ------------------------------
# 6. Create Program Start/End Rows
# ------------------------------
df_start = df_exploded.copy()
df_start["Program Start/End"] = "Program Start"
df_start["Time Analyzed"] = df_start["REQ_SCHED_START_DATE"]

df_end = df_exploded.copy()
df_end["Program Start/End"] = "Program End"
df_end["Time Analyzed"] = df_end["REQ_SCHED_END_DATE"]

df_long = pd.concat([df_start, df_end], ignore_index=True)

# ------------------------------
# 7. Merge with SAP + Fuzzy Date Match
# ------------------------------
df_joined = df_long.merge(sap_data, how="left", left_on="REQ_WO_NUM_CLEAN", right_on="Work Order #")
df_joined["Matched"] = (df_joined["Date"] - df_joined["Time Analyzed"]).abs() <= pd.Timedelta(days=2)

# Flag APP_NUM-level matches
match_summary = df_joined.groupby("APP_NUM")["Matched"].any().reset_index().rename(columns={"Matched": "App_Match_Found"})
df_final = df_joined.merge(match_summary, on="APP_NUM", how="left")

# ------------------------------
# 8. Final Filter: Time Within Analysis Window
# ------------------------------
df_final_filtered = df_final[df_final["Time Analyzed"].between(analysis_start, analysis_end)].copy()

# ------------------------------
# 9. Result
# ------------------------------
print("✅ Final Match Report:")
print(df_final_filtered[[
    "APP_NUM", "REQ_WO_NUM_CLEAN", "Program Start/End", "Time Analyzed",
    "Work Order #", "Date", "Matched", "App_Match_Found"
]])


## Update logic
# Step 1: Create Start and End filtered datasets
start_rows = df_exploded[df_exploded["REQ_SCHED_START_DATE"].between(analysis_start, analysis_end)].copy()
start_rows["Program Start/End"] = "Program Start"
start_rows["Time Analyzed"] = start_rows["REQ_SCHED_START_DATE"]

end_rows = df_exploded[df_exploded["REQ_SCHED_END_DATE"].between(analysis_start, analysis_end)].copy()
end_rows["Program Start/End"] = "Program End"
end_rows["Time Analyzed"] = end_rows["REQ_SCHED_END_DATE"]

# Step 2: Only keep these filtered rows for matching
df_long = pd.concat([start_rows, end_rows], ignore_index=True)

# Step 3: Merge with SAP
df_joined = df_long.merge(sap_data, how="left", left_on="REQ_WO_NUM_CLEAN", right_on="Work Order #")
df_joined["Matched"] = (df_joined["Date"] - df_joined["Time Analyzed"]).abs() <= pd.Timedelta(days=2)

# Step 4: Flag App_Match_Found at the APP_NUM level
match_summary = df_joined.groupby("APP_NUM")["Matched"].any().reset_index().rename(columns={"Matched": "App_Match_Found"})
df_final = df_joined.merge(match_summary, on="APP_NUM", how="left")

def extract_pure_sql(sas_code: str) -> str:
    lines = sas_code.splitlines()
    in_sql_block = False
    clean_lines = []

    for line in lines:
        stripped = line.strip()

        if stripped.lower().startswith("proc sql"):
            in_sql_block = True

        if in_sql_block:
            if not stripped.startswith("/*") and not stripped.startswith("%") and "ods " not in stripped.lower():
                clean_lines.append(line)

        if stripped.lower().startswith("quit;"):
            break

    return "\n".join(clean_lines)

## Even newer
import pandas as pd
import re

# ------------------------------
# 1. Setup: Input Data
# ------------------------------
itoa_data = pd.DataFrame({
    "APP_NUM": [1, 2, 3, 4, 5],
    "REQ_WO_NUM": [
        "TD11111",
        "TD22222, TD33333",
        "TD44444",
        "TD55555",
        "TD66666, TD77777, TD88888"
    ],
    "REQ_SCHED_START_DATE": pd.to_datetime([
        "2025-04-06", "2025-04-07", "2025-03-01", "2025-04-10", "2025-04-05"
    ]),
    "REQ_SCHED_END_DATE": pd.to_datetime([
        "2025-04-08", "2025-04-09", "2025-04-10", "2025-04-12", "2025-04-06"
    ])
})

sap_data = pd.DataFrame({
    "Work Order #": ["TD11111", "TD33333", "TD55555", "TD88888", "TD44444"],
    "Date": pd.to_datetime([
        "2025-04-07", "2025-04-08", "2025-04-11", "2025-04-06", "2025-04-10"
    ])
})

# ------------------------------
# 2. Config: Date Filter Window
# ------------------------------
analysis_start = pd.to_datetime("2025-04-06")
analysis_end = pd.to_datetime("2025-04-12")

# ------------------------------
# 3. Clean Work Order Numbers
# ------------------------------
def preprocess_wo_number(df):
    input_col = df["REQ_WO_NUM"].astype(str)
    df_output = pd.DataFrame()
    df_output["wo_original"] = input_col
    df_output["APP_NUM"] = df["APP_NUM"]

    df_output["removed_dashes"] = df_output["wo_original"].apply(
        lambda x: x.replace(",", "").replace("-", " ") if len(x.replace(",", "").split("-")[0]) < 9
        else x.replace(",", "").replace("-", "")
    )

    pattern = r"TD[ ]?\d{5,7}|90\d{7,8}|80\d{7,8}"
    df_output["all_wo"] = df_output["removed_dashes"].str.upper().apply(
        lambda x: re.findall(pattern, x)
    )

    return df_output

df_itoa_processed = preprocess_wo_number(itoa_data)
df_exploded = df_itoa_processed.explode("all_wo").drop_duplicates()
df_exploded = df_exploded.rename(columns={"all_wo": "REQ_WO_NUM_CLEAN"})

# ------------------------------
# 4. Reattach Scheduling Info
# ------------------------------
df_exploded = df_exploded.merge(
    itoa_data[["APP_NUM", "REQ_SCHED_START_DATE", "REQ_SCHED_END_DATE"]],
    on="APP_NUM", how="left"
)

# ------------------------------
# 5. Filter & Create Start/End Rows
# ------------------------------
start_rows = df_exploded[df_exploded["REQ_SCHED_START_DATE"].between(analysis_start, analysis_end)].copy()
start_rows["Program Start/End"] = "Program Start"
start_rows["Time Analyzed"] = start_rows["REQ_SCHED_START_DATE"]

end_rows = df_exploded[df_exploded["REQ_SCHED_END_DATE"].between(analysis_start, analysis_end)].copy()
end_rows["Program Start/End"] = "Program End"
end_rows["Time Analyzed"] = end_rows["REQ_SCHED_END_DATE"]

df_long = pd.concat([start_rows, end_rows], ignore_index=True)

# ------------------------------
# 6. Match with SAP (±2 Days)
# ------------------------------
df_joined = df_long.merge(sap_data, how="left", left_on="REQ_WO_NUM_CLEAN", right_on="Work Order #")
df_joined["Matched"] = (df_joined["Date"] - df_joined["Time Analyzed"]).abs() <= pd.Timedelta(days=2)
df_joined["Date_Matched"] = df_joined.apply(
    lambda row: row["Date"] if row["Matched"] else pd.NaT, axis=1
)

# ------------------------------
# 7. Flag APP_NUM-Level Match
# ------------------------------
match_summary = df_joined.groupby("APP_NUM")["Matched"].any().reset_index().rename(columns={"Matched": "App_Match_Found"})
df_final = df_joined.merge(match_summary, on="APP_NUM", how="left")

# ------------------------------
# 8. Final Output
# ------------------------------
df_final_sorted = df_final.sort_values(["APP_NUM", "Program Start/End", "REQ_WO_NUM_CLEAN"])
final_report = df_final_sorted[[
    "APP_NUM", "Program Start/End", "Time Analyzed",
    "REQ_WO_NUM_CLEAN", "Work Order #", "Date_Matched",
    "Matched", "App_Match_Found"
]]

# Print or export
print(final_report)
# You can also export to Excel: final_report.to_excel("itoa_sap_match_report.xlsx", index=False)




# 1. Merge
df_merged = df_long.merge(sap_data, how="left", left_on="REQ_WO_NUM_CLEAN", right_on="Work Order #")

# 2. Compute time delta
df_merged["TimeDelta"] = (df_merged["Date"] - df_merged["Time Analyzed"]).abs()

# 3. Filter to within 2-day window
df_windowed = df_merged[df_merged["TimeDelta"] <= pd.Timedelta(days=2)]

# 4. Keep only closest SAP match per row (based on APP_NUM + Program Start/End + WO)
deduped_idx = (
    df_windowed
    .groupby(["APP_NUM", "Program Start/End", "REQ_WO_NUM_CLEAN"])["TimeDelta"]
    .idxmin()
)

df_joined_best = df_windowed.loc[deduped_idx].copy()

# 5. Add match flags
df_joined_best["Matched"] = True
df_joined_best["Date_Matched"] = df_joined_best["Date"]

# 6. Fill in unmatched (i.e., rows that had no SAP match at all)
df_no_match = df_long.merge(
    df_joined_best[["APP_NUM", "Program Start/End", "REQ_WO_NUM_CLEAN"]],
    how="left",
    on=["APP_NUM", "Program Start/End", "REQ_WO_NUM_CLEAN"],
    indicator=True
).query("_merge == 'left_only'").drop(columns=["_merge"])

df_no_match["Matched"] = False
df_no_match["Date_Matched"] = pd.NaT

# 7. Final result
df_final = pd.concat([df_joined_best, df_no_match], ignore_index=True)






## EXPORT

import pandas as pd
import re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

# ------------------------------
# 2. Styled Excel Export Function
# ------------------------------
def export_styled_excel(df, filename="itoa_sap_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "iTOA-SAP Match"

    # Write the DataFrame to the Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Header styling
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font

    # Identify key columns
    headers = list(df.columns)
    match_idx = headers.index("Matched")
    prog_type_idx = headers.index("Program Start/End")
    sched_start_idx = headers.index("REQ_SCHED_START_DATE")
    sched_end_idx = headers.index("REQ_SCHED_END_DATE")

    # Row formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        match_cell = row[match_idx]
        prog_type = row[prog_type_idx].value

        # Highlight unmatched
        if match_cell.value is False:
            for cell in row:
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")

        # Dim irrelevant schedule column
        if prog_type == "Program Start":
            row[sched_end_idx].font = Font(color="999999")  # dim END
        elif prog_type == "Program End":
            row[sched_start_idx].font = Font(color="999999")  # dim START

    # Autofit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(filename)
    print(f"✅ Exported to {filename}")
    return filename








{
    "name": "Modern Dashboard Fusion",
    "dataColors": [
        "#FF2D75",
        "#FFD6E5",
        "#61d87c",
        "#54c6e9",
        "#f6db5e",
        "#e95474",
        "#3183c8",
        "#9b6be8",
        "#f28b74",
        "#dbdcdc"
    ],
    "background": "#F8F8F8",
    "foreground": "#000000",
    "tableAccent": "#FF2D75",
    "visualStyles": {
        "*": {
            "*": {
                "color": [
                    {
                        "solid": {
                            "color": "#000000"
                        }
                    }
                ],
                "background": [
                    {
                        "solid": {
                            "color": "#FFFFFF"
                        }
                    }
                ]
            }
        }
    },
    "sentimentColors": [
        "#61d87c",
        "#f6db5e",
        "#e95474"
    ]
}{
    "name": "Modern Dashboard Fusion",
    "dataColors": [
        "#FF2D75",
        "#FFD6E5",
        "#61d87c",
        "#54c6e9",
        "#f6db5e",
        "#e95474",
        "#3183c8",
        "#9b6be8",
        "#f28b74",
        "#dbdcdc"
    ],
    "background": "#F8F8F8",
    "foreground": "#000000",
    "tableAccent": "#FF2D75",
    "visualStyles": {
        "*": {
            "*": {
                "color": [
                    {
                        "solid": {
                            "color": "#000000"
                        }
                    }
                ],
                "background": [
                    {
                        "solid": {
                            "color": "#FFFFFF"
                        }
                    }
                ]
            }
        }
    },
    "sentimentColors": [
        "#61d87c",
        "#f6db5e",
        "#e95474"
    ]
}
def predict_folder(self,
                       img_filename,
                       # img_bytes,
                       model,
                       output_csv_file,
                       basePath):
        print(model._YoloEndpoint)
        print(model._PoleEndpoint)
        print(model._CrossarmEndpoint)
        print(model._TransformerEndpoint)
        print(model._InsulatorEndpoint)
        print(model._CotterKeyEndpoint)
        print(model._FuseEndpoint)
        print(model._SwitchBladeEndpoint)

        # modelMappingDict = {}
        # for item in self.flask_hyp['ROC_ML_PREDICTION_LIST']:
        #     try:
        #         with open(self.flask_hyp['CLASSIFIER_MAPPING_PATH'] + os.sep + "AERIAL_%s_DEFECT_MAPPING.json" % str(item).upper()) as infile:
        #             modelMappingDict[item] = json.load(infile)
        #     except Exception as e:
        #         print("No mapping file found for asset %s" % str(item))
        #         pass

        os.environ['TZ'] = 'America/Los_Angeles'
        file = ntpath.basename(img_filename)
        # outFrame = pd.DataFrame()
        durationList = []
        counter = 0
        try:
            start = time.time()
            # Added in forced rotations
            # imagesRotated = [cv2.cvtColor(cv2.imread(str(file)), cv2.COLOR_BGR2RGB)]
            # for idx, rotation in enumerate(imagesRotated):
            rotation = cv2.cvtColor(cv2.imread(str(file)), cv2.COLOR_BGR2RGB)
            try:
                object_process_start = time.time()
                currResult, _ = model.predict(imagePath=rotation,
                                              # img_bytes=img_bytes,
                                              genPlot=0,
                                              genPlotSavePath=basePath + os.sep + "PredImages" + os.sep + "%s_%s_pred.png" % (
                                                  str(
                                                      file).split(os.sep)[-1].split(".")[
                                                      0], str(0)),
                                              keywordFilter=[],
                                              confThresh=self._confThresh,  # .4 to .1
                                              probThresh=self._probThresh,  # 0.5 to 0.1
                                              vertex_serving=True
                                              )
                object_process_end = time.time()
                object_process_time = object_process_end - object_process_start
                print("Object Detection Prediction Time for file %s : %d seconds" % (file, int(object_process_time)))
                
                #Removing this - we want to store objects in DB for other use cases regardless of defect
                # ROC_ML_PREDICTION_LIST: ["POLE", "CROSSARM", "DEFECTIVE_POLE", "DEFECTIVE_CROSSARM"]
                # print(currResult)
                # currResult = currResult[currResult['label'].isin(model._ROC_ML_PREDICTION_LIST
                #                                                  )] if not currResult.empty else currResult
                if not currResult.empty:
                    currResult.reset_index(drop=True, inplace=True)
                    currResult['filename'] = str(img_filename)
                    currResult['rotation'] = 0
                    # if idx == 1:
                    #     currResult['rotation'] = 90
                    # if idx == 2:
                    #     currResult['rotation'] = -90
                    ROC_ML_PREDICTION_DICT = {
                        "POLE": model._PoleEndpoint,
                        "CROSSARM": model._CrossarmEndpoint,
                        "TRANSFORMER": model._TransformerEndpoint,
                        "INSULATOR": model._InsulatorEndpoint,
                        "COTTER_KEY": model._CotterKeyEndpoint,
                        "FUSE": model._FuseEndpoint,
                        "SWITCH_BLADE": model._SwitchBladeEndpoint,
                        "CAPACITOR": model._CapacitorEndpoint,
                        "LIGHTNING_ARRESTOR" : model._LightningArrestorEndpoint
                    }
                    print(list(ROC_ML_PREDICTION_DICT.keys()))
                    # Defect secondary classifier(s)
                    currResult['defect_probability'] = np.nan
                    currResult['defect_type'] = ""
                    currResult['priority_probability'] = np.nan
                    currResult['priority'] = ""
                    currResult['object_process_time'] = int(object_process_time)
                    currResult['defect_process_time'] = np.nan

                    #Force bird_nest defect probability to equal object probability for consistent output
                    if "BIRD_NEST" in currResult['label'].values.tolist():
                        currResult.loc[currResult['label']=='BIRD_NEST', 'defect_probability'] = currResult.loc[currResult['label']=='BIRD_NEST', 'object_probability']

                    #update loop logic to apply models in batches rather than 1 at a time(reduce time increase as we add new objects)
                    for asset in list(ROC_ML_PREDICTION_DICT.keys()):
                        print(asset)
                        asset_df = currResult[currResult['label'].str.contains(asset)]
                        if not asset_df.empty:
                            idxs_asset = asset_df.index
                            image = [preprocess_input(cv2.resize(rotation[ymin:ymax, xmin:xmax], (208, 208))) for
                                         xmin, xmax, ymin, ymax in
                                         asset_df[['xmin', 'xmax', 'ymin', 'ymax']].values.tolist()]

                            # image = np.expand_dims(preprocess_input(currCrop), 0)
                            image = np.array(image)
                            # print(image.shape)
                            # print(len(image.tolist()))
                            data = json.dumps({
                                "instances": image.tolist()
                            })
                            headers = {"content-type": "application/json"}
                            defect_process_start = time.time()
                            result = requests.post(
                                url=ROC_ML_PREDICTION_DICT[asset],
                                data=data, headers=headers)
                            defect_process_end = time.time()
                            defect_process_time = defect_process_end - defect_process_start
                            print(f"{asset} defect_process_time: {defect_process_time}")
                            pred = result.json()['predictions']

                            # objectDefect = pred[0]
                            # objectPriority = pred[1]

                            objectDefect = [x['defect_output'] for x in pred]
                            objectPriority = [x['severity_output'] for x in pred]

                            # print(objectDefect)
                            # if len(objectDefect) > 1:

                            priority = np.array(objectPriority)
                            defect = np.array(objectDefect)
                            keys = list(self.modelMappingDict[asset].keys())
                            vals = list(self.modelMappingDict[asset].values())

                            curr_keys = np.tile(keys, defect.shape[0]).reshape(defect.shape[0], -1)
                            # print(curr_keys)
                            # curr_defects = np.where(defect >= -1)
                            curr_defect_order = np.argsort(-defect)
                            curr_keys = np.take_along_axis(curr_keys, curr_defect_order, axis=-1)
                            # print(curr_keys)
                            defect = np.take_along_axis(defect, curr_defect_order, axis=-1)
                            defect = np.where(curr_keys == "NO_DEFECT", -1, defect)

                            #If all probabilities less than base value, take max for database
                            outputs_max_idx = np.argmax(defect, axis=-1)
                            # outputs_max_defect_type = curr_keys[outputs_max_idx]

                            outputs = np.where(defect >= .8, curr_keys, "NA")#.tolist()
                            # print(outputs)
                            outputs = [["@".join(filter(lambda x: x != "NA", y))] for y in outputs]

                            #if output is blank, use max value defect
                            outputs = [x if x != [''] else curr_keys[idx][outputs_max_idx[idx]] for idx, x in enumerate(outputs)]
                            #if a list exists, grab only the element
                            outputs = [x[0] if isinstance(x, list) else x for x in outputs]
                            #testing

                            #Adding priority prediction
                            priority_max_prob = np.argmax(priority, axis=-1)
                            priority_predictions_prob = priority[np.arange(priority.shape[0]), priority_max_prob]
                            priority_predictions = [self.priority_dict[x] for x in priority_max_prob]

                            currResult.loc[idxs_asset, 'defect_type'] = outputs
                            currResult.loc[idxs_asset, 'defect_probability'] = np.round((100 * np.max(defect, axis=-1))).astype(int).tolist()
                            currResult.loc[idxs_asset, 'priority_probability'] = np.round((100 * priority_predictions_prob)).astype(int)
                            currResult.loc[idxs_asset, 'priority'] = priority_predictions
                            # else:
                            #     currResult.loc[idxs_asset, 'defect_type'] = ""
                            #     defect = [x[0] for x in objectDefect]
                            #     # print(objectDefect)
                            #     currResult.loc[idxs_asset, 'defect_probability'] = np.round(100 * np.array(defect)).astype(int).tolist()
                            currResult.loc[idxs_asset, 'defect_process_time'] = defect_process_time
                        asset_df = pd.DataFrame()
                    # currResult = currResult[~currResult['defect_type'].str.contains("NO_DEFECT")]
                    context_logic_start = time.time()
                    currResult = self.context_logic.runLogic(df=currResult)
                    context_logic_end = time.time()
                    context_logic_time = context_logic_end - context_logic_start
                    print("context logic time: {} seconds".format(context_logic_time))

                    # for idx2, row2 in currResult.iterrows():
                    #     object = row2['label']
                    #     # for object in list(ROC_ML_PREDICTION_DICT.keys()):
                    #     # if object in row2['label']:
                    #     if object in list(ROC_ML_PREDICTION_DICT.keys()):
                    #         left = row2['xmin']
                    #         top = row2['ymin']
                    #         right = row2['xmax']
                    #         bottom = row2['ymax']
                    #
                    #         currCrop = rotation
                    #         currCrop = currCrop[top:bottom, left:right]
                    #         currCrop = cv2.resize(currCrop, (208, 208))
                    #
                    #         image = np.expand_dims(preprocess_input(currCrop), 0)
                    #         data = json.dumps({
                    #             "instances": image.tolist()
                    #         })
                    #         headers = {"content-type": "application/json"}
                    #
                    #         # creds, project = google.auth.default()
                    #         #
                    #         # # creds.valid is False, and creds.token is None
                    #         # # Need to refresh credentials to populate those
                    #         #
                    #         # auth_req = google.auth.transport.requests.Request()
                    #         # creds.refresh(auth_req)
                    #         # auth_token = creds.token
                    #         # headers = {"content-type": "application/json",
                    #         #            "Authorization": "Bearer %s" % auth_token}
                    #
                    #         result = requests.post(
                    #             url=ROC_ML_PREDICTION_DICT[object],
                    #             data=data, headers=headers)
                    #         objectDefect = result.json()['predictions']
                    #         objectDefect = objectDefect[0][0]
                    #         currResult.loc[idx2, 'defect_probability'] = round(100 * objectDefect)

                    currResult['filename'] = str(
                        img_filename)  # currResult['filename'].str)         #.replace("/", "\\")
                    currResult['prediction_datetime'] = str(MainUtil.get_pst_time())
                    currResult['xmin'] = currResult['xmin'] / currResult['img_width']
                    currResult['xmax'] = currResult['xmax'] / currResult['img_width']
                    currResult['ymin'] = currResult['ymin'] / currResult['img_height']
                    currResult['ymax'] = currResult['ymax'] / currResult['img_height']
                    currResult['label'] = currResult['label'].str.replace("DEFECTIVE_", "")
                else:
                    cols = ['filename', 'label', 'xmin', 'xmax', 'ymin', 'ymax', 'img_width', 'img_height',
                            'bboxcolor', 'label_timestamp', 'ignore', 'labeling_complete',
                            'rotation', 'object_probability', 'label_probability', 'defect_probability', 'defect_type',
                            'priority_probability', 'priority',
                            'prediction_datetime', 'object_process_time', 'defect_process_time']
                    currOut = [str(img_filename).replace("/", "\\"), "NO_PREDICTIONS", np.nan, np.nan, np.nan,
                               np.nan,
                               np.nan, np.nan, "NA", '3000-01-01_12:00:00', 1, 1, 0, np.nan, np.nan, np.nan,"", np.nan, "",
                               str(MainUtil.get_pst_time()), object_process_time, 0]
                    currResult = pd.DataFrame([dict(zip(cols, currOut))])
                    currResult = currResult[
                        ['filename', 'label', 'xmin', 'xmax', 'ymin', 'ymax', 'img_width', 'img_height',
                         'bboxcolor', 'label_timestamp', 'ignore', 'labeling_complete',
                         'rotation', 'object_probability', 'label_probability', 'defect_probability', 'defect_type',
                         'priority_probability', 'priority',
                         'prediction_datetime', 'object_process_time', 'defect_process_time']]
            except Exception as e:
                print(f'{e} for {img_filename}', flush=True)
                pass

            # temporary save after each iteration
            cols = ['filename', 'label', 'xmin', 'xmax', 'ymin', 'ymax', 'img_width', 'img_height',
                    'bboxcolor', 'label_timestamp', 'ignore', 'labeling_complete',
                    'rotation', 'object_probability', 'label_probability', 'defect_probability', 'defect_type',
                    'priority_probability', 'priority',
                    'prediction_datetime', 'object_process_time', 'defect_process_time']
            # output_csv_file = basePath + os.sep + str(pathOfInterest).split(os.sep)[-1].replace(":", "") + "_predictions.csv"
            # outFrame[cols].to_csv(output_csv_file, index=False)
            if not os.path.exists(output_csv_file):
                currResult[cols].to_csv(output_csv_file, header=True, index=False)
            else:
                currResult[cols].to_csv(output_csv_file, mode='a', header=False, index=False)

            currResult = None
            end = time.time()
            duration = (end - start) / 60
            durationList.append(duration)
            counter = counter + 1
            # status = float(counter) / float(fileNum)
            # timeRemaining = round((fileNum - counter) * np.mean(durationList), 2)
            # if counter % 10 == 0:
            #    print(str(round(status * 100, 2)) + "% Complete\nEstimated Time Remaining: {} Minutes".format(
            #        timeRemaining))
        except Exception as e:
            print(f'{e} for {img_filename}', flush=True)
            raise Exception(e)

        # print("Total run time: {}".format(round(np.sum(duration), 2)))
        # if not os.path.exists(basePath + "/PredImages"):
        #     os.mkdir(basePath + "/PredImages")
        #
        # outFrame.to_csv(basePath + "/PredImages/" + str(file).split("\\")[-1].replace(":", "") + "_predictions.csv", index=False)

        return (output_csv_file, False)
